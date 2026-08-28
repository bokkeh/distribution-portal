from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = ROOT / "knowledge-files"


def cell_value(value):
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


result = []
for path in sorted(SOURCE_DIR.glob("*.xlsx")):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for sheet in workbook.worksheets:
        max_row = sheet.max_row or 0
        max_column = sheet.max_column or 0
        preview = []
        if max_row > 0:
            for row in sheet.iter_rows(min_row=1, max_row=min(max_row, 12), values_only=True):
                preview.append([cell_value(value) for value in row])
        sheets.append(
            {
                "name": sheet.title,
                "max_row": max_row,
                "max_column": max_column,
                "preview": preview,
            }
        )
    result.append({"file": path.name, "sheets": sheets})

print(json.dumps(result, indent=2, ensure_ascii=False))
