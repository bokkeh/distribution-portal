from __future__ import annotations

import json
import math
import re
import unicodedata
from collections import defaultdict
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = ROOT / "knowledge-files"
CRM_STATE_PATH = ROOT / "tmp" / "crm_state.json"
OUTPUT_PATH = ROOT / "tmp" / "knowledge_audit.json"
IMPORT_TAG = "knowledge-audit-2026-08-27"

ORDER_FILES = [
    "Sales by Account order history 3_26_25.xlsx",
    "Sales by Account 042225.xlsx",
    "2025 Sales by Account.xlsx",
    "Sales Date per Account.xlsx",
]

NEW_FILES = {
    "Sales by Account order history 3_26_25.xlsx",
    "Sales by Account 042225.xlsx",
    "BBTT Tasting Data.xlsx",
    "Tasting Event statistics _ Where to find cocktail features.xlsx",
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text_value = str(value).replace("\ufffd", "'").replace("â€™", "'").replace("’", "'")
    return re.sub(r"\s+", " ", text_value).strip()


def normalize(value: Any) -> str:
    text_value = unicodedata.normalize("NFKD", clean_text(value).lower())
    text_value = "".join(char for char in text_value if not unicodedata.combining(char))
    text_value = text_value.replace("&", " and ")
    return re.sub(r"[^a-z0-9]+", " ", text_value).strip()


def iso_date(value: Any) -> str | None:
    parsed_date: date | None = None
    if isinstance(value, datetime):
        parsed_date = value.date()
    elif isinstance(value, date):
        parsed_date = value
    else:
        text_value = clean_text(value).lstrip("`")
        if not text_value or text_value == "-":
            return None
        for pattern in ("%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%m-%d-%y"):
            try:
                parsed_date = datetime.strptime(text_value, pattern).date()
                break
            except ValueError:
                continue
    if parsed_date is None or not 2000 <= parsed_date.year <= 2099:
        return None
    return parsed_date.isoformat()


def number(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return None
        return float(value)
    text_value = clean_text(value)
    if not text_value or text_value in {"-", "N/A", "n/a", "#DIV/0!"}:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text_value.replace(",", ""))
    return float(match.group()) if match else None


def integer(value: Any) -> int | None:
    parsed = number(value)
    return int(round(parsed)) if parsed is not None else None


def combine_source_notes(parts: list[tuple[str, Any]]) -> str | None:
    output = []
    for label, value in parts:
        text_value = clean_text(value)
        if text_value and text_value not in {"-", "N/A", "n/a", "None"}:
            output.append(f"{label}: {text_value}")
    return " | ".join(output) or None


def utc_timestamp(day: str, hour: int = 16, minute: int = 0) -> str:
    # Historical sources are Maryland/DC-local unless the matched account is in another state.
    # Noon/afternoon source values are stored at a stable UTC instant for account history ordering.
    parsed = datetime.fromisoformat(day).replace(hour=hour, minute=minute, tzinfo=timezone.utc)
    return parsed.isoformat().replace("+00:00", "Z")


def extract_time_parts(value: Any, event_day: str | None = None) -> tuple[int, int, int | None, int | None]:
    if isinstance(value, time):
        return value.hour, value.minute, None, None
    if isinstance(value, datetime):
        if value.hour or value.minute:
            return value.hour, value.minute, None, None
        if event_day and value.date().isoformat() != event_day and 1 <= value.month <= 12 and 1 <= value.day <= 12:
            start_hour = value.month
            end_hour = value.day
            if start_hour <= 7:
                start_hour += 12
            if end_hour <= 7:
                end_hour += 12
            return start_hour % 24, 0, end_hour % 24, 0
        return 16, 0, None, None
    if isinstance(value, (int, float)):
        text_value = str(int(value))
    else:
        text_value = clean_text(value).lower()
    if not text_value or text_value in {"-", "?", "???"}:
        return 16, 0, None, None
    nums = [int(part) for part in re.findall(r"\d+", text_value)]
    if not nums:
        return 16, 0, None, None

    def clock(raw: int) -> tuple[int, int]:
        if raw >= 100:
            hour_value, minute_value = divmod(raw, 100)
        else:
            hour_value, minute_value = raw, 0
        if hour_value <= 7:
            hour_value += 12
        return hour_value % 24, minute_value if minute_value < 60 else 0

    start_hour, start_minute = clock(nums[0])
    if len(nums) >= 2:
        end_hour, end_minute = clock(nums[1])
        return start_hour, start_minute, end_hour, end_minute
    return start_hour, start_minute, None, None


def timestamp_range(day: str, raw_start: Any, raw_end: Any = None) -> tuple[str, str]:
    start_hour, start_minute, embedded_end_hour, embedded_end_minute = extract_time_parts(raw_start, day)
    if raw_end is not None:
        end_hour, end_minute, _, _ = extract_time_parts(raw_end, day)
    elif embedded_end_hour is not None:
        end_hour, end_minute = embedded_end_hour, embedded_end_minute or 0
    else:
        end_hour, end_minute = (start_hour + 3) % 24, start_minute
    start = datetime.fromisoformat(day).replace(hour=start_hour, minute=start_minute, tzinfo=timezone.utc)
    end = datetime.fromisoformat(day).replace(hour=end_hour, minute=end_minute, tzinfo=timezone.utc)
    if end <= start:
        end += timedelta(days=1)
    return start.isoformat().replace("+00:00", "Z"), end.isoformat().replace("+00:00", "Z")


def localize_source_clock(timestamp_value: str, state: str | None) -> str:
    naive = datetime.fromisoformat(timestamp_value.replace("Z", "+00:00")).replace(tzinfo=None)
    year = naive.year

    def nth_weekday(month: int, weekday: int, occurrence: int) -> date:
        first = date(year, month, 1)
        days_until = (weekday - first.weekday()) % 7
        return first + timedelta(days=days_until + 7 * (occurrence - 1))

    dst_start = nth_weekday(3, 6, 2)
    dst_end = nth_weekday(11, 6, 1)
    is_dst = dst_start <= naive.date() < dst_end
    central = normalize(state) in {"il", "illinois"}
    offset_hours = (-5 if is_dst else -6) if central else (-4 if is_dst else -5)
    localized = naive.replace(tzinfo=timezone(timedelta(hours=offset_hours)))
    return localized.astimezone(timezone.utc).isoformat().replace("+00:00", "Z")


def parse_orders() -> list[dict[str, Any]]:
    output = []
    state_labels = {
        "maryland", "washington dc", "dc", "florida", "illinois", "grand totals",
        "maryland total", "dc total", "florida total", "illinois total",
    }
    for file_name in ORDER_FILES:
        workbook = load_workbook(SOURCE_DIR / file_name, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            current_account = ""
            current_state = ""
            for row_number, row in enumerate(sheet.iter_rows(values_only=True), 1):
                values = list(row) + [None] * max(0, 16 - len(row))
                raw_label = clean_text(values[0])
                normalized_label = normalize(raw_label).replace(" totals", " total")
                if normalized_label in state_labels or normalized_label.endswith(" total"):
                    if normalized_label in {"maryland", "washington dc", "dc", "florida", "illinois"}:
                        current_state = raw_label
                    continue
                if raw_label:
                    current_account = raw_label
                order_day = iso_date(values[3])
                cases = number(values[2])
                if not current_account or not order_day or cases is None or cases <= 0:
                    continue
                output.append({
                    "kind": "order",
                    "source_file": file_name,
                    "source_sheet": sheet.title,
                    "source_row": row_number,
                    "is_new_file": file_name in NEW_FILES,
                    "state_section": current_state,
                    "source_account": current_account,
                    "order_date": order_day,
                    "delivery_date": iso_date(values[4]),
                    "paid_date": iso_date(values[5]),
                    "cases": cases,
                    "source_total": number(values[6]),
                    "source_notes": clean_text(values[14]) or None,
                })
    return output


def parse_bbtt_tastings() -> list[dict[str, Any]]:
    file_name = "BBTT Tasting Data.xlsx"
    workbook = load_workbook(SOURCE_DIR / file_name, read_only=True, data_only=True)
    output = []
    for sheet in workbook.worksheets:
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), 1):
            values = list(row) + [None] * max(0, 23 - len(row))
            activity_type = clean_text(values[0])
            day = iso_date(values[3])
            location = clean_text(values[6])
            if row_number == 1 or not activity_type or not day or not location:
                continue
            scheduled_at, end_at = timestamp_range(day, values[4], values[5])
            output.append({
                "kind": "tasting",
                "source_priority": 1,
                "source_file": file_name,
                "source_sheet": sheet.title,
                "source_row": row_number,
                "is_new_file": True,
                "activity_type": activity_type,
                "source_account": location,
                "date": day,
                "scheduled_at": scheduled_at,
                "end_at": end_at,
                "source_email": clean_text(values[1]) or None,
                "source_taster": clean_text(values[7]) or None,
                "samples_served": integer(values[16]),
                "bottles_sold": integer(values[18]),
                "consumer_interactions": integer(values[15]),
                "bottles_in_stock_before": integer(values[10]),
                "bottle_price_on_shelf": number(values[11]),
                "account_feedback": clean_text(values[20]) or None,
                "highlights": combine_source_notes([
                    ("Cocktail sampled", values[13]),
                    ("Notable comments", values[19]),
                ]),
                "issues": combine_source_notes([
                    ("Weather/factors", values[8]),
                    ("Signage", values[9]),
                    ("Low materials", values[21]),
                    ("Issues/questions", values[22]),
                ]),
                "source_notes": combine_source_notes([
                    ("Activity type", values[0]),
                    ("Total traffic", values[14]),
                    ("Purchased by team member", values[17]),
                    ("Tasting price", values[12]),
                ]),
            })
    return output


def parse_legacy_tastings() -> list[dict[str, Any]]:
    file_name = "Tasting Event statistics _ Where to find cocktail features.xlsx"
    workbook = load_workbook(SOURCE_DIR / file_name, read_only=True, data_only=True)
    output = []
    sheet_rules = {
        "MDDC Liquor store": {"header": 2, "location": 0, "date": 3, "time": 4, "taster": 5, "cocktail": 7, "stock": 8, "notes": 9, "traffic": 10, "talked": 11, "tasted": 12, "sold": 13, "price": 17},
        "FL Liquor Store": {"header": 1, "location": 0, "date": 3, "time": 4, "taster": 5, "cocktail": 7, "stock": 8, "notes": 9, "traffic": 10, "talked": 11, "tasted": 12, "sold": 13, "price": 16},
        "Restaurants": {"header": 1, "location": 0, "date": 3, "time": 4, "taster": None, "cocktail": 19, "stock": 5, "notes": 6, "traffic": 7, "talked": None, "tasted": 8, "sold": 9, "price": 12},
    }
    for sheet_name, rule in sheet_rules.items():
        sheet = workbook[sheet_name]
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if row_number <= rule["header"]:
                continue
            values = list(row)
            day = iso_date(values[rule["date"]] if len(values) > rule["date"] else None)
            location = clean_text(values[rule["location"]] if len(values) > rule["location"] else None)
            if not day or not location:
                continue
            raw_time = values[rule["time"]] if len(values) > rule["time"] else None
            scheduled_at, end_at = timestamp_range(day, raw_time)
            notes_value = values[rule["notes"]] if rule["notes"] is not None and len(values) > rule["notes"] else None
            row_context = " ".join(clean_text(value) for value in values if value is not None)
            status = "cancelled" if re.search(r"\b(cxl|cancel|cancelled|canceled)\w*\b", row_context, re.I) else "completed"
            cancellation_context = None
            if status == "cancelled":
                cancellation_parts = [
                    clean_text(value)
                    for value in values
                    if isinstance(value, str) and re.search(r"\b(cxl|cancel|cancelled|canceled)\w*\b", clean_text(value), re.I)
                ]
                cancellation_context = " | ".join(dict.fromkeys(cancellation_parts)) or None

            def at(key: str) -> Any:
                index = rule[key]
                return values[index] if index is not None and len(values) > index else None

            output.append({
                "kind": "tasting",
                "source_priority": 2,
                "source_file": file_name,
                "source_sheet": sheet_name,
                "source_row": row_number,
                "is_new_file": True,
                "activity_type": "Tasting" if sheet_name != "Restaurants" else "Restaurant activity",
                "source_account": location,
                "date": day,
                "scheduled_at": scheduled_at,
                "end_at": end_at,
                "source_email": None,
                "source_taster": clean_text(at("taster")) or None,
                "status": status,
                "samples_served": integer(at("tasted")),
                "bottles_sold": integer(at("sold")),
                "consumer_interactions": integer(at("talked")) or integer(at("traffic")),
                "bottles_in_stock_before": integer(at("stock")),
                "bottle_price_on_shelf": number(at("price")),
                "account_feedback": None,
                "highlights": combine_source_notes([("Cocktail sampled", at("cocktail"))]),
                "issues": cancellation_context or clean_text(notes_value) or None,
                "source_notes": combine_source_notes([
                    ("Legacy source sheet", sheet_name),
                    ("Traffic", at("traffic")),
                ]),
            })
    return output


ACCOUNT_ALIASES = {
    normalize("-GAD Promo Cases"): "Global Alliance Distribution- PO",
    normalize("1 West Dupont"): "1 West Dupont Circle Wines and Liquors",
    normalize("Amendment 21"): "Amendment 21",
    normalize("Aspen Hill"): "Aspen Hill Montgomery County ABS MoCo Liquor store",
    normalize("Barstool River North"): "Barstool River North",
    normalize("Cabin John"): "Cabin John Montgomery County MoCo ABS liquor store",
    normalize("Castaways Beach Club"): "Castaways Beach Club",
    normalize("Cattail Creek Country Club"): "Cattail Creek Country Club",
    normalize("Christo's Discount Liquors"): "Christos Discount Liquors",
    normalize("Christos Discount Liquors"): "Christos Discount Liquors",
    normalize("Clarksburg Village"): "Clarksburg Village",
    normalize("Clarksburg"): "Clarksburg Montgomery County MoCo ABS Liquor Store",
    normalize("MoCo ABS Clarksburg"): "Clarksburg Montgomery County MoCo ABS Liquor Store",
    normalize("MoCo Oak Barrel & Vine @ Clarksburg"): "Clarksburg Montgomery County MoCo ABS Liquor Store",
    normalize("Cloverly"): "Cloverly Montgomery County ABS MoCo",
    normalize("Darnestown"): "Darnestown Montgomery County ABS MoCo",
    normalize("Darnstown"): "Darnestown Montgomery County ABS MoCo",
    normalize("Downtown Rockville"): "Downtown Rockville Montgomery County ABS MoCo",
    normalize("MoCo Downtown Rockville"): "Downtown Rockville Montgomery County ABS MoCo",
    normalize("MoCo Oak Barrel & Vine @ Downtown Rockville"): "Downtown Rockville Montgomery County ABS MoCo",
    normalize("Eastport"): "Eastport Liquors",
    normalize("Eastport Liquors"): "Eastport Liquors",
    normalize("Fallsgrove"): "Fallsgrove Montgomery County ABS MoCo",
    normalize("Shady Grove"): "Fallsgrove Montgomery County ABS MoCo",
    normalize("Flower"): "Flower Montgomery County ABS Moco",
    normalize("MoCo Oak Barrel & Vine @ Flower"): "Flower Montgomery County ABS Moco",
    normalize("Gaithersburg Square"): "Gaithersburg Square Montgomery County ABS MoCo Store",
    normalize("Galliano's"): "Galliano's Itailian Restaurant and Wine Bar",
    normalize("Galliano"): "Galliano's Itailian Restaurant and Wine Bar",
    normalize("Global Alliance Distribution- PO"): "Global Alliance Distribution- PO",
    normalize("Glenwood Liquors"): "glenwood wine and spirits",
    normalize("Glenwood Wine & Spirits"): "glenwood wine and spirits",
    normalize("Glenwood Wines"): "glenwood wine and spirits",
    normalize("Goshen"): "Goshen Crossing Montgomery County ABS MoCo",
    normalize("Goshen Crossing"): "Goshen Crossing Montgomery County ABS MoCo",
    normalize("MoCo ABS Goshen"): "Goshen Crossing Montgomery County ABS MoCo",
    normalize("Hampden Lane"): "Hampden Lane Montgomery County ABS MoCo",
    normalize("Highland Wine and Spirits"): "Highland Wine and spirits",
    normalize("Hotel Zena"): "Hotel Zena",
    normalize("Kensington"): "Kensington Montgomery Country Liquor & Wine",
    normalize("King Farm (MoCo)"): "King Farm Montgomery (MoCo) Liquor",
    normalize("Kings Farm"): "King Farm Montgomery (MoCo) Liquor",
    normalize("Moco King Farm"): "King Farm Montgomery (MoCo) Liquor",
    normalize("MoCo Liquor & Wine at Kings Farm"): "King Farm Montgomery (MoCo) Liquor",
    normalize("Montgomery County Kings Farm"): "King Farm Montgomery (MoCo) Liquor",
    normalize("Gaithersburg (Kings Farm Village Center)"): "King Farm Montgomery (MoCo) Liquor",
    normalize("Kingsview"): "Kingsview Montgomery County Liquor ABS MoCo",
    normalize("Kingsview (MoCo)"): "Kingsview Montgomery County Liquor ABS MoCo",
    normalize("MoCo Kingsview"): "Kingsview Montgomery County Liquor ABS MoCo",
    normalize("Konstantine's Greek Taverna"): "Konstantine's Greek Taverna",
    normalize("La Caj Seafood Restaurant"): "La Caj Seafood Restaurant",
    normalize("Leisure World"): "Leisure World Montgomery County ABS MoCo",
    normalize("MoCo ABS Leisure World"): "Leisure World Montgomery County ABS MoCo",
    normalize("Liquor Expo"): "Liquor Expo",
    normalize("Moco ABS Muddy Branch"): "Muddy Branch Montgomery county MoCo ABS",
    normalize("Muddy Branch (MoC0)"): "Muddy Branch (MoC0)",
    normalize("Muddy Branch"): "Muddy Branch Montgomery county MoCo ABS",
    normalize("Montrose"): "Montrose Montgomery County ABS MoCo",
    normalize("MoCo ABS Montrose"): "Montrose Montgomery County ABS MoCo",
    normalize("Old Line Kitchen"): "Old Line Kitchen",
    normalize("Old Tyme"): "Olde Tyme Liquors",
    normalize("Olde Tyme"): "Olde Tyme Liquors",
    normalize("Olde Tyme Liquors"): "Olde Tyme Liquors",
    normalize("Tasting Olde Tyme Liquors"): "Olde Tyme Liquors",
    normalize("Olde tyme liqours"): "Olde Tyme Liquors",
    normalize("Olney"): "Olney Wine and Liquor Montgomery County Moco ABS",
    normalize("MoCo ABS Olney"): "Olney Wine and Liquor Montgomery County Moco ABS",
    normalize("Periodic Table"): "Periodic Table",
    normalize("Perfect Pour"): "The Perfect Pour",
    normalize("The Perfect Pour"): "The Perfect Pour",
    normalize("Perry's Restaurant"): "Perry's Restaurant",
    normalize("Petite Cellar"): "Petite Cellars",
    normalize("Petite Cellars"): "Petite Cellars",
    normalize("Poolesville"): "Poolesville Montgomery County ABS MoCo",
    normalize("Montgomery County Liquor & Wine @ Poolesville"): "Poolesville Montgomery County ABS MoCo",
    normalize("Potomac"): "Potomac Montgomery County ABS Moco",
    normalize("MoCo ABS Potomac"): "Potomac Montgomery County ABS Moco",
    normalize("Renaissance Baltimore Harborplace"): "Renaissance Baltimore Harborplace",
    normalize("Seneca Meadows"): "Seneca Meadows Montgomery County Moco Liquor ABS",
    normalize("Seneca Meadows (MoCo)"): "Seneca Meadows Montgomery County Moco Liquor ABS",
    normalize("MoCo Seneca Meadows"): "Seneca Meadows Montgomery County Moco Liquor ABS",
    normalize("Silver Spring"): "Silver Spring Montgomery County ABS MoCo",
    normalize("MoCo ABS Silver Spring"): "Silver Spring Montgomery County ABS MoCo",
    normalize("Montgomery County Liquor & Wine Silver Spring"): "Silver Spring Montgomery County ABS MoCo",
    normalize("Snowden Liquors"): "Snowden River Liquors",
    normalize("Snowden River Liquors"): "Snowden River Liquors",
    normalize("Snowden River Liquor"): "Snowden River Liquors",
    normalize("Snowden"): "Snowden River Liquors",
    normalize("Star Liquors"): "Star Liquors",
    normalize("Tino's"): "Tino's Itailan Bistro and Wine Bar",
    normalize("Total Discount Liquors"): "Total Discount Liquors",
    normalize("Total Discount Liquor"): "Total Discount Liquors",
    normalize("Total Discount Liquors, Eldersburg"): "Total Discount Liquors",
    normalize("Walnut Hill"): "Walnut Hill Montgomery County MoCo ABS",
    normalize("Walnut Hill (MoCo)"): "Walnut Hill Montgomery County MoCo ABS",
    normalize("MoCo ABS Walnut Hill"): "Walnut Hill Montgomery County MoCo ABS",
    normalize("Moco ABS Walnut HIll"): "Walnut Hill Montgomery County MoCo ABS",
    normalize("Wheaton"): "Wheaton Montgomery Count ABS MoCo",
    normalize("White Oak"): "White Oak Montgomery County ABS MoCo",
    normalize("MoCo ABS White Oak"): "White Oak Montgomery County ABS MoCo",
    normalize("Wunder Garten"): "Wunder Garten",
    normalize("MoCo ABS Aspen Hill"): "Aspen Hill Montgomery County ABS MoCo Liquor store",
    normalize("MoCo Oak Barrel & Vine @ Aspen Hill"): "Aspen Hill Montgomery County ABS MoCo Liquor store",
    normalize("MoCo ABS Gaithersburg"): "Gaithersburg Square Montgomery County ABS MoCo Store",
    normalize("MoCo ABS Kensington"): "Kensington Montgomery Country Liquor & Wine",
    normalize("Northridge Wine and Spirits"): "North Ridge wine and spirits",
    normalize("Olde Tyme Liquiors"): "Olde Tyme Liquors",
    normalize("Perfect Pour, Elkridge"): "The Perfect Pour",
    normalize("Sunny's Fine Wine and Spirits"): "Sunnyâ€™s Fine Wines and Liquors",
    normalize("Sunnys Time Wines & Liquor"): "Sunnyâ€™s Fine Wines and Liquors",
    normalize("Hotel Zena Meta meet up"): "Hotel Zena",
}


USER_NAME_ALIASES = {
    "andrew": "Andrew Murray",
    "andrew murray": "Andrew Murray",
    "andrew murphy": "Andrew Murray",
    "connor": "Connor McIntyre",
    "connor mcintyre": "Connor McIntyre",
    "connor and andrew": "Connor McIntyre",
    "kim": "Kimberly LaRose (admin)",
    "kim and andrew murray": "Kimberly LaRose (admin)",
    "kim and emily": "Kimberly LaRose (admin)",
    "kim and natalie": "Kimberly LaRose (admin)",
    "kim subbed for jay": "Kimberly LaRose (admin)",
    "kimberly larose": "Kimberly LaRose (admin)",
    "kimberly larose keyah": "Kimberly LaRose (admin)",
    "kimberly larose sean": "Kimberly LaRose (admin)",
    "kimberly larose sales": "Kimberly LaRose (admin)",
    "kimberly larose admin": "Kimberly LaRose (admin)",
    "natalie": "Natalie Heinrichs",
    "natalie heinrichs": "Natalie Heinrichs",
    "rachel anderson": "Rachel Anderson",
    "sean": "Sean Jordan",
    "sean jordan": "Sean Jordan",
}


def main() -> None:
    crm = json.loads(CRM_STATE_PATH.read_text(encoding="utf-8"))
    accounts_by_name = {account["companyName"]: account for account in crm["accounts"]}
    accounts_by_normalized = defaultdict(list)
    for account in crm["accounts"]:
        accounts_by_normalized[normalize(account["companyName"])].append(account)
    users_by_email = {normalize(user["email"]): user for user in crm["users"]}
    users_by_name = {user["name"]: user for user in crm["users"]}

    def match_account(source_account: str) -> tuple[dict[str, Any] | None, str]:
        normalized = normalize(source_account)
        alias_target = ACCOUNT_ALIASES.get(normalized)
        if alias_target:
            direct = accounts_by_name.get(alias_target)
            if direct:
                return direct, "alias"
            normalized_targets = accounts_by_normalized.get(normalize(alias_target), [])
            if len(normalized_targets) == 1:
                return normalized_targets[0], "alias"
        exact = accounts_by_normalized.get(normalized, [])
        if len(exact) == 1:
            return exact[0], "exact"
        return None, "unmatched"

    def match_user(row: dict[str, Any]) -> tuple[dict[str, Any] | None, str]:
        email_key = normalize(row.get("source_email"))
        if email_key and email_key in users_by_email:
            return users_by_email[email_key], "email"
        name_key = normalize(row.get("source_taster"))
        target_name = USER_NAME_ALIASES.get(name_key)
        if target_name and target_name in users_by_name:
            return users_by_name[target_name], "alias"
        return None, "unmatched"

    order_rows = parse_orders()
    for row in order_rows:
        account, method = match_account(row["source_account"])
        row["account_match_method"] = method
        row["account_id"] = account["id"] if account else None
        row["account_name"] = account["companyName"] if account else None
        row["candidate_key"] = f'{row["account_id"]}:{row["order_date"]}:{row["cases"]:.2f}' if account else None

    existing_order_keys = set()
    for row in crm["orders"]:
        day = row["createdAt"][:10]
        quantity = number(row.get("quantity"))
        if quantity is not None:
            existing_order_keys.add(f'{row["customerId"]}:{day}:{quantity:.2f}')

    # Deduplicate repeated snapshots, preferring the latest/most comprehensive workbook order.
    order_file_priority = {name: index for index, name in enumerate(ORDER_FILES)}
    best_orders: dict[str, dict[str, Any]] = {}
    for row in order_rows:
        key = row.get("candidate_key")
        if not key:
            continue
        current = best_orders.get(key)
        if current is None or order_file_priority[row["source_file"]] > order_file_priority[current["source_file"]]:
            best_orders[key] = row
    order_candidates = []
    for key, row in best_orders.items():
        row["disposition"] = "existing" if key in existing_order_keys else "insert"
        if row["source_file"] in NEW_FILES or row["disposition"] == "insert":
            order_candidates.append(row)

    tasting_rows = parse_bbtt_tastings() + parse_legacy_tastings()
    for row in tasting_rows:
        account, account_method = match_account(row["source_account"])
        user, user_method = match_user(row)
        row["account_match_method"] = account_method
        row["account_id"] = account["id"] if account else None
        row["account_name"] = account["companyName"] if account else None
        if account:
            row["scheduled_at"] = localize_source_clock(row["scheduled_at"], account.get("state"))
            row["end_at"] = localize_source_clock(row["end_at"], account.get("state"))
        row["user_match_method"] = user_method
        row["assigned_user_id"] = user["id"] if user else None
        row["assigned_user_name"] = user["name"] if user else None
        row.setdefault("status", "completed")
        if (
            row["source_file"] == "Tasting Event statistics _ Where to find cocktail features.xlsx"
            and row["source_sheet"] == "MDDC Liquor store"
            and row["source_row"] == 82
        ):
            row["data_anomaly_reason"] = "Source date is 2025-10-12 inside a chronological block of 2024 rows."
        elif row["date"] > "2025-12-31":
            row["data_anomaly_reason"] = "Source tasting date falls outside the workbook's historical 2022-2025 range."

    # Source-to-source deduplication: account + calendar date, preferring cleaned BBTT rows.
    tasting_groups = defaultdict(list)
    for row in tasting_rows:
        if row["account_id"]:
            tasting_groups[(row["account_id"], row["date"])].append(row)
    deduped_tastings = []
    source_duplicates = []
    for key, rows in tasting_groups.items():
        sorted_rows = sorted(
            rows,
            key=lambda item: (
                item["source_priority"],
                0 if item["assigned_user_id"] else 1,
                0 if item["samples_served"] is not None else 1,
            ),
        )
        chosen = sorted_rows[0]
        deduped_tastings.append(chosen)
        source_duplicates.extend(sorted_rows[1:])

    existing_tasting_days = {(row["customerId"], row["scheduledAt"][:10]) for row in crm["tastings"]}
    for row in deduped_tastings:
        if row.get("data_anomaly_reason"):
            row["disposition"] = "skip_data_anomaly"
        elif not row["assigned_user_id"]:
            row["disposition"] = "skip_missing_taster"
        elif (row["account_id"], row["date"]) in existing_tasting_days:
            row["disposition"] = "existing"
        else:
            row["disposition"] = "insert"

    unmatched_tastings = [row for row in tasting_rows if not row["account_id"]]
    matched_missing_taster = [row for row in deduped_tastings if row["disposition"] == "skip_missing_taster"]
    tasting_anomalies = [row for row in deduped_tastings if row["disposition"] == "skip_data_anomaly"]
    insert_tastings = [row for row in deduped_tastings if row["disposition"] == "insert"]
    existing_tastings = [row for row in deduped_tastings if row["disposition"] == "existing"]
    unmatched_orders = [row for row in order_rows if not row["account_id"] and row["source_file"] in NEW_FILES]
    insert_orders = [row for row in order_candidates if row["disposition"] == "insert"]

    audit = {
        "import_tag": IMPORT_TAG,
        "summary": {
            "source_order_rows": len(order_rows),
            "new_file_order_rows": sum(1 for row in order_rows if row["source_file"] in NEW_FILES),
            "order_insert_candidates": len(insert_orders),
            "new_file_unmatched_order_rows": len(unmatched_orders),
            "source_tasting_rows": len(tasting_rows),
            "matched_unique_tasting_rows": len(deduped_tastings),
            "tasting_insert_candidates": len(insert_tastings),
            "existing_tasting_rows": len(existing_tastings),
            "matched_missing_taster_rows": len(matched_missing_taster),
            "tasting_anomaly_rows": len(tasting_anomalies),
            "unmatched_tasting_rows": len(unmatched_tastings),
            "source_duplicate_tasting_rows": len(source_duplicates),
        },
        "order_insert_candidates": insert_orders,
        "order_candidates_reviewed": order_candidates,
        "unmatched_order_rows": unmatched_orders,
        "tasting_insert_candidates": insert_tastings,
        "existing_tasting_rows": existing_tastings,
        "matched_missing_taster_rows": matched_missing_taster,
        "tasting_anomaly_rows": tasting_anomalies,
        "unmatched_tasting_rows": unmatched_tastings,
        "source_duplicate_tasting_rows": source_duplicates,
    }
    OUTPUT_PATH.write_text(json.dumps(audit, indent=2, ensure_ascii=False), encoding="utf-8")
    print(json.dumps({"summary": audit["summary"], "output": str(OUTPUT_PATH.relative_to(ROOT))}, indent=2))


if __name__ == "__main__":
    main()
